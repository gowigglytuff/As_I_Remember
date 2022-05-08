from openpyxl import load_workbook


class GameSpreadsheet(object):
    def __init__(self, gc_input, gd_input):
        self.gc_input = gc_input
        self.gd_input = gd_input
        self.file = None
        self.name =None
        self.sheet_names_list = []

    def spreadsheet_get_phrase(self, character_name):
        gc_input_time = self.gc_input.time_of_day
        gc_input_day = self.gc_input.day_of_summer
        gc_input_outfit = self.gd_input.player["Player"].current_outfit
        return_phrase = "Hey thanks!"

        ss_accessed = self.access_spreadsheet(character_name)
        ss_time = ss_accessed[0]
        indexes = ss_accessed[1]

        for row in ss_time:
            flag1 = False
            flag2 = False
            flag3 = False
            all_flags_true = False
            # time
            if row[indexes[0]] == gc_input_time:
                flag1 = True

            # date
            if row[indexes[1]] == gc_input_day:
                flag2 = True

            # outfit
            if row[indexes[2]] == gc_input_outfit:
                flag3 = True

            if flag1 and flag2 and flag3 == True:
                all_flags_true = True

            if all_flags_true:
                return_phrase = row['dialogue']
                print(row['dialogue'])

        return return_phrase

    def access_spreadsheet(self, character_name):
        ss_time = self.load_spreadsheet(file_name=self.file, sheet_name=character_name)

        ss_keys = list(ss_time[0].keys())

        indexes = [k for k in ss_keys if ss_keys.index(k) < 3]

        return ss_time, indexes


    def load_spreadsheet(self, file_name, sheet_name):
        workbook = load_workbook(filename=file_name)

        sheet = workbook[sheet_name]
        # for sheet in workbook.worksheets:

        row_count = sheet.max_row
        column_count = sheet.max_column

        header = []
        for col in range(1, column_count + 1):
            header.append(sheet.cell(1, col).value)

        row_values = []

        for row in range(2, row_count + 1):
            this_row = {}
            for ind, header_name in enumerate(header):
                this_row[header_name] = sheet.cell(row, ind+1).value
            row_values.append(this_row)


        for row in row_values:
            pass
            # print(row)
        return row_values


class ThanksSpreadsheet(GameSpreadsheet):
    NAME = "Thanks"
    def __init__(self, gc_input, gd_input):
        super().__init__(gc_input, gd_input)
        self.file = "assets/spreadsheets/thanks_sheet.xlsx"
        self.name = self.NAME
        self.workbook = load_workbook(self.file)
        self.sheet_names_list = self.workbook.sheetnames
        print(self.sheet_names_list)

    def spreadsheet_get_phrase(self, character_name, reaction):
        return_phrase = 'gggzzzgGGzzz eerrooorr-'

        if character_name in self.sheet_names_list:
            ss_accessed = self.access_spreadsheet(character_name)
        else:
            ss_accessed = self.access_spreadsheet("Generic")
        ss_time = ss_accessed[0]
        indexes = ss_accessed[1]

        index_of_interest = indexes.index(reaction)

        result = "default"
        for row in ss_time:
            if row[indexes[index_of_interest]] == 1:
                result = ss_time.index(row)

        if result != "default":
            return_phrase = ss_time[result]['dialogue']
            print(return_phrase)


        return return_phrase


class PlayerLocationSheet(GameSpreadsheet):
    NAME = "player_location"

    def __init__(self, gc_input, gd_input):
        super().__init__(gc_input, gd_input)
        self.file = "assets/spreadsheets/Player_location.xlsx"
        self.name = self.NAME
        self.workbook = load_workbook(self.file)
        self.sheet_names_list = self.workbook.sheetnames


    def spreadsheet_load_location(self):
        ss_accessed = self.load_spreadsheet()
        SS_player = ss_accessed[0]

        return SS_player

    def get_indexes(self):

        ss_time = self.load_spreadsheet(file_name=self.file, sheet_name="Player")

        ss_keys = list(ss_time[0].keys())

        return ss_time

    def load_spreadsheet(self):
        workbook = load_workbook(filename=self.file)

        sheet = workbook["Player"]
        # for sheet in workbook.worksheets:

        row_count = sheet.max_row
        column_count = sheet.max_column
        print(row_count, column_count)

        header = []
        for col in range(1, column_count + 1):
            header.append(sheet.cell(1, col).value)

        row_values = []

        for row in range(2, row_count + 1):
            this_row = {}
            for ind, header_name in enumerate(header):
                this_row[header_name] = sheet.cell(row, ind+1).value
            row_values.append(this_row)

        return row_values

    def write_to_workbook(self, player_x, player_y, camera_x, camera_y, current_room):
        sheet = self.workbook["Player"]
        column_count = sheet.max_column

        for column in range(column_count):
            print(column)
            column += 1
            if self.workbook["Player"].cell(1, column).value == "player_x":
                self.workbook["Player"].cell(2, column).value = player_x
            if self.workbook["Player"].cell(1, column).value == "player_y":
                self.workbook["Player"].cell(2, column).value = player_y
            if self.workbook["Player"].cell(1, column).value == "camera_x":
                self.workbook["Player"].cell(2, column).value = camera_x
            if self.workbook["Player"].cell(1, column).value == "camera_y":
                self.workbook["Player"].cell(2, column).value = camera_y
            if self.workbook["Player"].cell(1, column).value == "current_room":
                self.workbook["Player"].cell(2, column).value = current_room

        self.workbook.save(self.file)